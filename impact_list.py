from jnpr.junos import Device
import time
from openpyxl import load_workbook

switch_data="D:\\amr_ali\Switches Database - 2017 (UPDATED).xlsx"

wb = load_workbook(switch_data)
sheet1=wb.get_sheet_by_name("SWs in Production")
#print (sheet1.cell(row=716,column=9).value)
sw_ip=input("please enter SWITCH ip without spaces  : ")
for i in range(1,1000):
    if sheet1.cell(row=i,column=14).value == sw_ip:
        print (  "Switch name is : "+sheet1.cell(row=i,column=13).value +"\n"
                 + " Region : "+sheet1.cell(row=i,column=7).value+"\n"
                 + "CID is  : "+ sheet1.cell(row=i,column=22).value+"\n")


